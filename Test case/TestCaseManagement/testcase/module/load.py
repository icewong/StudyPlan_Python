#-*- coding=utf-8 -*-

from django.core.exceptions import ObjectDoesNotExist
from ..MACRO import *
import os

from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from ..models import project,project_path,AllCases
def UpdateTestCasesFromFile(path,project_id):
    try:
        this_pro = project.objects.get(id=project_id)
    except ObjectDoesNotExist:
        return


    if os.path.exists(path):
        src_wb = load_workbook(filename=path, read_only=True)
        src_sheet = src_wb['测试用例']

        vaild_data=False

        for row in src_sheet.rows:
            if not vaild_data:
                if row[0].value.strip()=='*Testcase ID':
                    vaild_data=True
                    continue
            else:
                valid_cell=[cell.value for index,cell in enumerate(row) if index<7]
                #jump if content is None, except (*pre condition)
                need_jump=False
                for index,item in enumerate(valid_cell):
                    if valid_cell[index]==None:
                        if index==4:
                            valid_cell[index]=''
                        else:
                            need_jump=True
                            break
                if need_jump:
                    continue
                path_items=[i for i in valid_cell[2].split('/') if len(i)>0]
                parent_index=0
                item_path=''
                case_path_item =None
                for item in path_items:
                    try:
                        case_path_item = project_path.objects.get(project=this_pro,name=item,parent_id=parent_index)
                        parent_index = case_path_item.id
                    except ObjectDoesNotExist:
                        case_path_item = project_path(name=item,parent_id=parent_index,content=item_path,project=this_pro)
                        case_path_item.save()
                        parent_index = case_path_item.id
                    item_path+='/'+item

                cur_index = valid_cell[0]
                cur_type = valid_cell[1]
                cur_name = valid_cell[3]
                cur_pre_condition = valid_cell[4]
                cur_step = valid_cell[5]
                cur_exception = valid_cell[6]

                try:
                    res = AllCases.objects.filter(path=case_path_item).get(case_index=cur_index)
                    res.test_type = cur_type
                    res.name = cur_name
                    res.pre_condition =cur_pre_condition
                    res.step = cur_step
                    res.exception = cur_exception
                    res.save()
                except ObjectDoesNotExist:
                    case=AllCases(case_index=cur_index,
                                  test_type=cur_type,
                                  name=cur_name,
                                  pre_condition=cur_pre_condition,
                                  step=cur_step,
                                  exception=cur_exception,
                                  path=case_path_item)
                    case.save()
        src_wb.close()
    else:
        print('Can not find file!!!!')

from ..models import instance_cases
from ..Functions import InstanceRelated
from openpyxl import Workbook
from openpyxl.styles import Side,NamedStyle,Border,PatternFill,Font
def ExportTestcase(instance_id):

    def append_cells_with_style(work_sheet,append_cells,style):
        work_sheet.append(append_cells)
        for col in range(1,work_sheet.max_column+1):
            work_sheet.cell(row=work_sheet.max_row,column=col).style=style

    try:
        dst_book = Workbook()
        dst_sheet = dst_book.get_active_sheet()
        dst_sheet.title='TestCases'

        #style
        title_style = NamedStyle(name="title")
        title_style.fill=PatternFill(start_color='FFCCFFFF',fill_type='solid')
        title_style.font=Font(b=True)
        bd=Side(style='thin',color='000000')
        title_style.border=Border(left=bd, right=bd, top=bd, bottom=bd)
        title_style.alignment.vertical='center'

        title_style2 = NamedStyle(name="title2")
        title_style2.font=Font(b=True)
        title_style2.fill=PatternFill(start_color='FFFCD5B4',fill_type='solid')
        title_style2.border=Border(left=bd, right=bd, top=bd, bottom=bd)
        title_style2.alignment.vertical='center'

        normal_style = NamedStyle(name="normal")
        normal_style.border=Border(left=bd, right=bd, top=bd, bottom=bd)
        normal_style.alignment.wrap_text = True

        title_items=['*Testcase ID','*Test Type','Testitem Path','*Testcase Name','*pre condition','*Steps',\
                         '*Expection','*Result',]

        dst_sheet.append(title_items)
        for col in range(1,9):
            dst_sheet.cell(row=dst_sheet.max_row, column=col).style=title_style
        for col in range(9,10):
            dst_sheet.cell(row=dst_sheet.max_row, column=col).style=title_style2

        case_col_widths=[15,15,30,20,20,20,20,10]
        for index,col in enumerate(dst_sheet.columns):
            if index==len(case_col_widths):
                break
            column = col[0].column
            dst_sheet.column_dimensions[column].width = case_col_widths[index]

        query_ins_case = instance_cases.objects.filter(instance_id=instance_id)
        for item in query_ins_case:
            case_item= item.case
            res_cells = [case_item.case_index,
                         case_item.test_type,
                         case_item.path.content+'/'+case_item.path.name+'/',
                         case_item.name,
                         case_item.pre_condition,
                         case_item.step,
                         case_item.exception,
                         translate_result(item.result)]
            append_cells_with_style(dst_sheet, res_cells, normal_style)
        return save_virtual_workbook(dst_book)
    except ObjectDoesNotExist:
        return []